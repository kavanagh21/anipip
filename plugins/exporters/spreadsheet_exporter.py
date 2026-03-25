"""Spreadsheet exporter plugin — writes TableData to CSV or Excel."""

import csv
from pathlib import Path
from typing import Callable

from core.plugin_base import BasePlugin
from core.image_container import ImageContainer
from core.table_data import TableData
from core.pipeline_data import PipelineData
from core.ports import InputPort, OutputPort, Port
from core.parameters import BoolParameter, ChoiceParameter, FileParameter, StringParameter


class SpreadsheetExporter(BasePlugin):
    """Write measurement / tabular data to CSV or Excel.

    In **batch mode** (the default when running a batch pipeline), rows are
    accumulated via :meth:`batch_initialize` / :meth:`batch_finalize` and
    written once at the end.  In **single mode**, the file is written
    immediately on each execution.

    Connect the optional *Image In* port to obtain the source image
    folder for *Save to Source Image Folder*.

    For Excel output you can specify a sheet name.  When appending to
    an existing sheet whose columns do not match the incoming data, a
    new sheet is created with ``(clash)`` appended to the name.
    """

    name = "Spreadsheet Exporter"
    category = "Exporters"
    description = "Export tabular measurement data to CSV or Excel"
    help_text = (
        "Writes measurement table data to a CSV or Excel (.xlsx) file. "
        "Enable \"Save to Source Image Folder\" and connect the optional "
        "Image In port to write alongside the original images. Set a "
        "custom filename to control the output name. For Excel, specify "
        "a sheet name \u2014 if the sheet has existing data with different "
        "columns, a new sheet with \"(clash)\" is created. In batch mode, "
        "rows are accumulated and written once at the end."
    )
    icon = None

    # -- Ports ---------------------------------------------------------

    ports: list[Port] = [
        InputPort("measurements", TableData, label="Measurements"),
        InputPort("image_in", ImageContainer, label="Image In", optional=True),
        OutputPort("table_out", TableData, label="Table Out"),
    ]

    # -- Parameters ----------------------------------------------------

    parameters = [
        BoolParameter(
            name="save_to_source_folder",
            label="Save to Source Image Folder",
            default=False,
        ),
        BoolParameter(
            name="use_output_subfolder",
            label="Save into 'output' subfolder",
            default=False,
        ),
        ChoiceParameter(
            name="format",
            label="Format",
            choices=["CSV", "Excel (.xlsx)"],
            default="CSV",
        ),
        FileParameter(
            name="output_folder",
            label="Output Folder",
            default="",
            folder_mode=True,
        ),
        StringParameter(
            name="custom_filename",
            label="Custom Filename",
            default="",
            placeholder="e.g. results (without extension)",
        ),
        StringParameter(
            name="sheet_name",
            label="Excel Sheet Name",
            default="Measurements",
            placeholder="e.g. Colocalization",
        ),
        BoolParameter(
            name="show_in_viewer",
            label="Show in internal viewer",
            default=True,
        ),
    ]

    def __init__(self):
        super().__init__()
        self._accumulated: TableData | None = None
        self._last_table: TableData | None = None
        self._source_folder: Path | None = None

    # -- Batch lifecycle -----------------------------------------------

    def batch_initialize(self) -> None:
        self._accumulated = TableData()
        self._source_folder = None

    def batch_finalize(self) -> None:
        if self._accumulated and self._accumulated.rows:
            path = self._resolve_output_path()
            if path:
                self._write_table(self._accumulated, path)
            self._last_table = self._accumulated
        self._accumulated = None

    # -- Processing ----------------------------------------------------

    def process(
        self,
        image: ImageContainer,
        progress_callback: Callable[[float], None],
    ) -> ImageContainer:
        return image

    def process_ports(
        self,
        inputs: dict[str, PipelineData],
        progress_callback: Callable[[float], None],
    ) -> dict[str, PipelineData]:
        progress_callback(0.1)

        table: TableData | None = inputs.get("measurements")
        if table is None:
            progress_callback(1.0)
            return {}

        # Capture source folder from any connected ImageContainer input
        for inp in inputs.values():
            if isinstance(inp, ImageContainer) and inp.metadata.source_path:
                self._source_folder = inp.metadata.source_path.parent
                break

        if self._accumulated is not None:
            self._accumulated = self._accumulated.merge(table)
        else:
            path = self._resolve_output_path()
            if path:
                self._write_table(table, path)

        self._last_table = table

        progress_callback(1.0)

        if self.get_parameter("show_in_viewer"):
            out = self._accumulated if self._accumulated is not None else table
            return {"table_out": out}

        return {}

    # -- Path resolution -----------------------------------------------

    def _resolve_output_path(self) -> Path | None:
        """Build the full output file path from the current settings."""
        save_to_source = self.get_parameter("save_to_source_folder")
        custom_name = (self.get_parameter("custom_filename") or "").strip()
        fmt = self.get_parameter("format")

        ext = ".xlsx" if fmt == "Excel (.xlsx)" else ".csv"

        # Strip extension if the user included one
        if custom_name:
            p = Path(custom_name)
            if p.suffix.lower() in (".csv", ".xlsx", ".xls"):
                custom_name = p.stem

        filename = f"{custom_name}{ext}" if custom_name else f"measurements{ext}"

        # Determine folder
        folder: Path | None = None

        if save_to_source:
            # Try local _source_folder first, then pipeline-level fallback
            sf = self._source_folder or self._pipeline_source_folder
            if sf and sf.exists():
                folder = sf

        if folder is None:
            folder_str = (self.get_parameter("output_folder") or "").strip()
            if folder_str:
                folder = Path(folder_str)

        if folder is None:
            return None

        if self.get_parameter("use_output_subfolder"):
            folder = folder / "output"

        folder.mkdir(parents=True, exist_ok=True)
        return folder / filename

    # -- Writing -------------------------------------------------------

    def _write_table(self, table: TableData, output_path: Path) -> None:
        fmt = self.get_parameter("format")
        output_path.parent.mkdir(parents=True, exist_ok=True)

        if fmt == "Excel (.xlsx)":
            self._write_xlsx(table, output_path)
        else:
            self._write_csv(table, output_path)

    @staticmethod
    def _write_csv(table: TableData, output_path: Path) -> None:
        if output_path.exists():
            # Append rows
            with open(output_path, "a", newline="") as f:
                writer = csv.DictWriter(
                    f, fieldnames=table.columns, extrasaction="ignore"
                )
                if output_path.stat().st_size == 0:
                    writer.writeheader()
                for row in table.rows:
                    writer.writerow(row)
        else:
            with open(output_path, "w", newline="") as f:
                writer = csv.DictWriter(
                    f, fieldnames=table.columns, extrasaction="ignore"
                )
                writer.writeheader()
                for row in table.rows:
                    writer.writerow(row)

    def _write_xlsx(self, table: TableData, output_path: Path) -> None:
        try:
            import openpyxl
        except ImportError:
            raise RuntimeError(
                "Excel export requires the openpyxl package. "
                "Install it with: pip install openpyxl"
            )

        sheet_name = (self.get_parameter("sheet_name") or "").strip()
        if not sheet_name:
            sheet_name = "Measurements"

        if output_path.exists():
            wb = openpyxl.load_workbook(output_path)
        else:
            wb = openpyxl.Workbook()
            # Remove the default empty sheet — we'll create our own
            if wb.sheetnames == ["Sheet"]:
                del wb["Sheet"]

        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # Check if existing columns match incoming columns
            if ws.max_row and ws.max_row >= 1:
                existing_cols = [
                    ws.cell(row=1, column=c).value
                    for c in range(1, ws.max_column + 1)
                ]
                # Strip None values from trailing empty columns
                existing_cols = [
                    c for c in existing_cols if c is not None
                ]

                if existing_cols and existing_cols != table.columns:
                    # Column mismatch — create a new sheet with (clash)
                    clash_name = f"{sheet_name} (clash)"
                    # Find a unique name if (clash) already exists
                    counter = 1
                    while clash_name in wb.sheetnames:
                        counter += 1
                        clash_name = f"{sheet_name} (clash {counter})"

                    ws = wb.create_sheet(title=clash_name)
                    # Write header
                    for col_idx, col_name in enumerate(
                        table.columns, start=1
                    ):
                        ws.cell(row=1, column=col_idx, value=col_name)
                # else: columns match — append below
        else:
            # Sheet doesn't exist yet — create it with header
            ws = wb.create_sheet(title=sheet_name)
            for col_idx, col_name in enumerate(table.columns, start=1):
                ws.cell(row=1, column=col_idx, value=col_name)

        # Append data rows
        start_row = ws.max_row + 1
        for row_idx, row_data in enumerate(table.rows, start=start_row):
            for col_idx, col_name in enumerate(table.columns, start=1):
                ws.cell(
                    row=row_idx, column=col_idx,
                    value=row_data.get(col_name, ""),
                )

        wb.save(output_path)

    # -- Validation ----------------------------------------------------

    def validate_parameters(self) -> tuple[bool, list[str]]:
        errors = []
        save_to_source = self.get_parameter("save_to_source_folder")
        output_folder = (self.get_parameter("output_folder") or "").strip()
        show_in_viewer = self.get_parameter("show_in_viewer")

        if not save_to_source and not output_folder and not show_in_viewer:
            errors.append(
                "Specify an output folder, enable 'Save to Source Image "
                "Folder', or enable 'Show in internal viewer'"
            )

        return len(errors) == 0, errors
